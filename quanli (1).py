import os
import sqlite3
from abc import ABC, abstractmethod
from dataclasses import dataclass
from tkinter import Tk, Frame, Label, Entry, Button, StringVar, IntVar, messagebox, ttk, filedialog, Scrollbar, VERTICAL, HORIZONTAL, RIGHT, BOTTOM, Y, X, BOTH
import openpyxl
try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None


DB_FILE = "thu_vien.db"
IMAGE_MENU = "library.png"
IMAGE_BG = "image2.png"
IMAGE_LOGIN = "finance.png"

# =========================
# THEME - MAU GIAO DIEN HIEN DAI
# =========================
COLOR_PANEL = "#06111F"
COLOR_PANEL_2 = "#0B1B2F"
COLOR_ACCENT = "#3B82F6"
COLOR_ACCENT_HOVER = "#60A5FA"
COLOR_TEXT = "#EAF2FF"
COLOR_MUTED = "#9DB3D1"
COLOR_INPUT = "#10243D"
COLOR_TABLE = "#0B1628"


# =========================
# MODEL - CAC LOP DOI TUONG
# =========================
# Phan nay duoc viet lai de the hien day du 4 tinh chat OOP:
# 1. Dong goi: du lieu duoc an trong cac thuoc tinh __private, truy cap qua @property.
# 2. Ke thua: Book ke thua LibraryItem, BorrowRecord ke thua LibraryTransaction.
# 3. Da hinh: cac lop con cung cai dat to_db_tuple() nhung tra ve du lieu khac nhau.
# 4. Truu tuong: LibraryItem va LibraryTransaction la lop truu tuong, bat buoc lop con cai dat method.

class DatabaseObject(ABC):
    """Lop truu tuong chung cho cac doi tuong co the chuyen thanh du lieu luu CSDL."""

    @abstractmethod
    def to_db_tuple(self):
        """Moi lop con phai tu dinh nghia cach chuyen doi tuong thanh tuple de luu database."""
        pass


class LibraryItem(DatabaseObject):
    """Lop cha truu tuong cho cac tai lieu trong thu vien."""

    def __init__(self, item_id, title):
        # Dong goi: khong cho truy cap truc tiep tu ben ngoai bang item.__item_id
        self.__item_id = str(item_id).strip().upper()
        self.__title = str(title).strip().title()

    @property
    def item_id(self):
        return self.__item_id

    @property
    def title(self):
        return self.__title

    def is_valid_base_info(self):
        """Kiem tra thong tin chung cua moi tai lieu."""
        return bool(self.__item_id and self.__title)


class Book(LibraryItem):
    """Lop Book ke thua LibraryItem, mo rong them tac gia, the loai, so luong, vi tri."""

    def __init__(self, book_id, title, author, genre, copies, location):
        super().__init__(book_id, title)
        self.__author = str(author).strip().title()
        self.__genre = str(genre).strip().title()
        self.__copies = int(copies)
        self.__location = str(location).strip().title()

    # Cac property duoc dung de doc du lieu an ben trong lop
    @property
    def book_id(self):
        return self.item_id

    @property
    def author(self):
        return self.__author

    @property
    def genre(self):
        return self.__genre

    @property
    def copies(self):
        return self.__copies

    @property
    def location(self):
        return self.__location

    def is_valid(self):
        """Kiem tra day du thong tin sach."""
        return self.is_valid_base_info() and bool(self.__author and self.__genre and self.__location) and self.__copies >= 0

    def to_db_tuple(self):
        """Da hinh: Book tu dinh nghia cach chuyen doi tuong thanh tuple de them vao bang book_info."""
        return (self.book_id, self.title, self.author, self.genre, self.copies, self.location)


class LibraryTransaction(DatabaseObject):
    """Lop cha truu tuong cho cac giao dich trong thu vien."""

    def __init__(self, book_id, student_id):
        self.__book_id = str(book_id).strip().upper()
        self.__student_id = str(student_id).strip().upper()

    @property
    def book_id(self):
        return self.__book_id

    @property
    def student_id(self):
        return self.__student_id

    def is_valid_transaction(self):
        return bool(self.__book_id and self.__student_id)


class BorrowRecord(LibraryTransaction):
    """Lop BorrowRecord ke thua LibraryTransaction, bieu dien mot lan muon sach."""

    def __init__(self, book_id, student_id, issue_date="", return_date=""):
        super().__init__(book_id, student_id)
        self.__issue_date = issue_date
        self.__return_date = return_date

    @property
    def issue_date(self):
        return self.__issue_date

    @property
    def return_date(self):
        return self.__return_date

    def to_db_tuple(self):
        """Da hinh: BorrowRecord cung co to_db_tuple() nhung phuc vu bang book_issued."""
        return (self.book_id, self.student_id)


# =========================
# DATABASE - LOP XU LY CSDL
# =========================
class LibraryDatabase:
    def __init__(self, db_file=DB_FILE):
        self.db_file = db_file
        self.create_tables()

    def connect(self):
        return sqlite3.connect(self.db_file)

    def create_tables(self):
        with self.connect() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS book_info (
                    ID TEXT PRIMARY KEY NOT NULL,
                    TITLE TEXT NOT NULL,
                    AUTHOR TEXT NOT NULL,
                    GENRE TEXT NOT NULL,
                    COPIES INTEGER NOT NULL,
                    LOCATION TEXT NOT NULL
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS book_issued (
                    BOOK_ID TEXT NOT NULL,
                    STUDENT_ID TEXT NOT NULL,
                    ISSUE_DATE TEXT NOT NULL,
                    RETURN_DATE TEXT NOT NULL,
                    PRIMARY KEY (BOOK_ID, STUDENT_ID)
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS login (
                    mem_id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT,
                    username TEXT,
                    password TEXT
                )
            """)
            cur = conn.execute("SELECT * FROM login WHERE username = ? AND password = ?", ("admin", "admin"))
            if cur.fetchone() is None:
                conn.execute("INSERT INTO login(username, password) VALUES(?, ?)", ("admin", "admin"))

    def check_login(self, username, password):
        with self.connect() as conn:
            cur = conn.execute(
                "SELECT * FROM login WHERE username = ? AND password = ?",
                (username, password)
            )
            return cur.fetchone() is not None

    def add_book(self, book: Book):
        with self.connect() as conn:
            # Da hinh: LibraryDatabase khong can biet chi tiet ben trong Book,
            # chi can goi to_db_tuple() cua doi tuong duoc truyen vao.
            conn.execute(
                "INSERT INTO book_info VALUES (?, ?, ?, ?, ?, ?)",
                book.to_db_tuple(),
            )

    def import_books_from_excel(self, file_path):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Chua cai thu vien openpyxl. Hay chay lenh: pip install openpyxl")

        workbook = load_workbook(file_path)
        sheet = workbook.active
        added_count = 0
        skipped_count = 0

        with self.connect() as conn:
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                try:
                    book_id = str(row[0]).strip().upper()
                    title = str(row[1]).strip().title()
                    author = str(row[2]).strip().title()
                    genre = str(row[3]).strip().title()
                    copies = int(row[4])
                    location = str(row[5]).strip().title()

                    if not all([book_id, title, author, genre, location]) or copies < 0:
                        skipped_count += 1
                        continue

                    conn.execute(
                        "INSERT INTO book_info VALUES (?, ?, ?, ?, ?, ?)",
                        (book_id, title, author, genre, copies, location),
                    )
                    added_count += 1
                except Exception:
                    skipped_count += 1

        return added_count, skipped_count

    def search_books(self, keyword):
        keyword = keyword.strip().title()
        keyword_id = keyword.upper()
        with self.connect() as conn:
            cur = conn.execute(
                """
                SELECT * FROM book_info
                WHERE ID = ? OR TITLE = ? OR AUTHOR = ? OR GENRE = ?
                """,
                (keyword_id, keyword, keyword, keyword),
            )
            return cur.fetchall()

    def get_all_books(self):
        with self.connect() as conn:
            cur = conn.execute("SELECT * FROM book_info ORDER BY TITLE")
            return cur.fetchall()

    def add_copies(self, book_id, number):
        with self.connect() as conn:
            conn.execute("UPDATE book_info SET COPIES = COPIES + ? WHERE ID = ?", (number, book_id))

    def delete_copies(self, book_id, number):
        with self.connect() as conn:
            cur = conn.execute("SELECT COPIES FROM book_info WHERE ID = ?", (book_id,))
            row = cur.fetchone()
            if row is None:
                raise ValueError("Khong tim thay sach")
            if number > row[0]:
                raise ValueError("So luong xoa lon hon so sach hien co")
            conn.execute("UPDATE book_info SET COPIES = COPIES - ? WHERE ID = ?", (number, book_id))

    def delete_book(self, book_id):
        with self.connect() as conn:
            cur = conn.execute("SELECT * FROM book_issued WHERE BOOK_ID = ?", (book_id,))
            if cur.fetchall():
                raise ValueError("Sach dang duoc muon, khong the xoa")
            conn.execute("DELETE FROM book_info WHERE ID = ?", (book_id,))

    def issue_book(self, record: BorrowRecord):
        book_id = record.book_id
        student_id = record.student_id
        with self.connect() as conn:
            cur = conn.execute("SELECT COPIES FROM book_info WHERE ID = ?", (book_id,))
            row = cur.fetchone()
            if row is None:
                raise ValueError("Ma sach khong ton tai")
            if row[0] <= 0:
                raise ValueError("Sach da het, khong the muon")
            # Da hinh: BorrowRecord tu cung cap du lieu can luu thong qua to_db_tuple().
            conn.execute(
                "INSERT INTO book_issued VALUES (?, ?, date('now'), date('now', '+7 day'))",
                record.to_db_tuple(),
            )
            conn.execute("UPDATE book_info SET COPIES = COPIES - 1 WHERE ID = ?", (book_id,))

    def return_book(self, book_id, student_id):
        book_id = book_id.upper()
        student_id = student_id.upper()
        with self.connect() as conn:
            cur = conn.execute(
                "SELECT * FROM book_issued WHERE BOOK_ID = ? AND STUDENT_ID = ?",
                (book_id, student_id),
            )
            if cur.fetchone() is None:
                raise ValueError("Khong tim thay du lieu muon sach")
            conn.execute(
                "DELETE FROM book_issued WHERE BOOK_ID = ? AND STUDENT_ID = ?",
                (book_id, student_id),
            )
            conn.execute("UPDATE book_info SET COPIES = COPIES + 1 WHERE ID = ?", (book_id,))

    def search_activity(self, keyword):
        keyword = keyword.strip().upper()
        with self.connect() as conn:
            cur = conn.execute(
                "SELECT * FROM book_issued WHERE BOOK_ID = ? OR STUDENT_ID = ?",
                (keyword, keyword),
            )
            return cur.fetchall()

    def get_all_activity(self):
        with self.connect() as conn:
            cur = conn.execute("SELECT * FROM book_issued ORDER BY ISSUE_DATE DESC")
            return cur.fetchall()


# =========================
# VIEW - GIAO DIEN TKINTER
# =========================
class LibraryApp:
    def __init__(self):
        self.db = LibraryDatabase()
        self.root = Tk()
        self.root.title("Dang nhap he thong thu vien")
        self.root.state("zoomed")
        self.configure_styles()
        self.current_frame = None
        self.active_panel = None
        self.bg_image_ref = None
        self.show_login()
        self.root.mainloop()

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def configure_styles(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.configure(
            "Treeview",
            background=COLOR_TABLE,
            foreground=COLOR_TEXT,
            fieldbackground=COLOR_TABLE,
            rowheight=30,
            font=("Segoe UI", 10),
            borderwidth=0,
        )
        style.configure(
            "Treeview.Heading",
            background=COLOR_PANEL_2,
            foreground=COLOR_ACCENT_HOVER,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
        )
        style.map("Treeview", background=[("selected", COLOR_ACCENT)], foreground=[("selected", "white")])

    def center_position(self, width, height, y_shift=0):
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = max((sw - width) // 2, 20)
        y = max((sh - height) // 2 + y_shift, 20)
        return x, y

    def make_label(self, parent, text, size=13, bold=True, fg=COLOR_TEXT, bg=COLOR_PANEL):
        weight = "bold" if bold else "normal"
        return Label(parent, text=text, font=("Segoe UI", size, weight), bg=bg, fg=fg)

    def make_button(self, parent, text, command=None, width=18, size=13, bg=COLOR_PANEL_2, fg=COLOR_TEXT):
        return Button(
            parent,
            text=text,
            font=("Segoe UI", size, "bold"),
            bg=bg,
            fg=fg,
            activebackground=COLOR_ACCENT,
            activeforeground="white",
            relief="flat",
            bd=0,
            width=width,
            cursor="hand2",
            command=command,
        )

    def make_entry(self, parent, variable, width=30, show=None):
        return Entry(
            parent,
            textvariable=variable,
            show=show,
            font=("Segoe UI", 12),
            bg=COLOR_INPUT,
            fg=COLOR_TEXT,
            insertbackground=COLOR_TEXT,
            relief="flat",
            width=width,
        )

    def create_background(self, image_path):
        self.clear_window()
        w = self.root.winfo_screenwidth()
        h = self.root.winfo_screenheight()
        frame = Frame(self.root, bg=COLOR_PANEL)
        frame.pack(fill="both", expand=True)

        if Image and ImageTk and os.path.exists(image_path):
            img = Image.open(image_path)
            img = img.resize((w, h), Image.Resampling.LANCZOS)
            self.bg_image_ref = ImageTk.PhotoImage(img)
            bg = Label(frame, image=self.bg_image_ref)
            bg.place(x=0, y=0, relwidth=1, relheight=1)
        return frame

    def make_panel(self, parent, width=650, height=520, x=None, y=None):
        # Xoa khung lam viec cu de cac chuc nang hien trong cung 1 menu, khong bi chong giao dien.
        if self.active_panel is not None and self.active_panel.winfo_exists():
            self.active_panel.destroy()

        if x is None or y is None:
            x, y = self.center_position(width, height)

        panel = Frame(parent, bg=COLOR_PANEL, width=width, height=height, highlightthickness=1, highlightbackground="#1D3B63")
        panel.place(x=x, y=y)
        panel.pack_propagate(False)
        self.active_panel = panel
        return panel

    def show_login(self):
        # Dung nen login rieng. Khung dang nhap duoc dich sang phai
        # de khong che mat chu "QUAN LY THU VIEN" tren anh nen.
        canvas = self.create_background(IMAGE_LOGIN if os.path.exists(IMAGE_LOGIN) else IMAGE_BG)
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        panel_width = 460
        panel_height = 390
        panel_x = int(sw * 0.7) - panel_width // 2
        panel_y = (sh - panel_height) // 2
        panel = self.make_panel(canvas, width=panel_width, height=panel_height, x=panel_x, y=panel_y)

        self.make_label(panel, "DANG NHAP ADMIN", size=24, fg=COLOR_ACCENT_HOVER).pack(pady=(28, 22))

        self.username_var = StringVar()
        self.password_var = StringVar()

        self.make_label(panel, "Ten dang nhap", size=12, fg=COLOR_MUTED).pack(anchor="w", padx=62)
        self.make_entry(panel, self.username_var, width=32).pack(pady=(6, 16), ipady=8)

        self.make_label(panel, "Mat khau", size=12, fg=COLOR_MUTED).pack(anchor="w", padx=62)
        self.make_entry(panel, self.password_var, width=32, show="*").pack(pady=(6, 22), ipady=8)

        self.make_button(panel, "DANG NHAP", command=self.login, width=24, size=12, bg=COLOR_ACCENT, fg="white").pack(pady=(0, 18), ipady=6)
        self.make_label(panel, "Tai khoan mac dinh: admin / admin", size=10, bold=False, fg=COLOR_MUTED).pack()

    def login(self):
        username = self.username_var.get().strip()
        password = self.password_var.get().strip()
        if not username or not password:
            messagebox.showwarning("Thong bao", "Vui long nhap day du ten dang nhap va mat khau")
            return
        if self.db.check_login(username, password):
            self.show_main_menu()
        else:
            messagebox.showerror("Loi", "Ten dang nhap hoac mat khau khong dung")

    def show_main_menu(self):
        """Menu chinh gom ca quan ly sach va muon tra sach trong cung 1 man hinh."""
        canvas = self.create_background(IMAGE_BG)
        self.active_panel = None

        title_w, title_h = 560, 66
        title_x, _ = self.center_position(title_w, title_h, y_shift=-310)
        Label(
            canvas,
            text="HE THONG QUAN LY THU VIEN",
            font=("Segoe UI", 24, "bold"),
            bg=COLOR_PANEL,
            fg=COLOR_ACCENT_HOVER,
            padx=20,
            pady=10,
            highlightthickness=1,
            highlightbackground="#1D3B63",
        ).place(x=title_x, y=25, width=title_w, height=title_h)

        menu_x = 25
        self.make_label(canvas, "QUAN LY SACH", size=15, fg=COLOR_ACCENT_HOVER, bg=COLOR_PANEL).place(x=menu_x, y=90, width=260, height=42)
        self.make_button(canvas, "Them sach", self.show_add_book, width=20, size=14).place(x=menu_x, y=150, width=260, height=48)
        self.make_button(canvas, "Tim sach", self.show_search_book, width=20, size=14).place(x=menu_x, y=215, width=260, height=48)
        self.make_button(canvas, "Tat ca sach", self.show_all_books, width=20, size=14).place(x=menu_x, y=280, width=260, height=48)
        self.make_button(canvas, "Doc tu Excel", self.import_books_excel, width=20, size=14).place(x=menu_x, y=345, width=260, height=48)

        self.make_label(canvas, "MUON TRA SACH", size=15, fg=COLOR_ACCENT_HOVER, bg=COLOR_PANEL).place(x=menu_x, y=435, width=260, height=42)
        self.make_button(canvas, "Muon sach", self.show_issue_book, width=20, size=14).place(x=menu_x, y=495, width=260, height=48)
        self.make_button(canvas, "Tra sach", self.show_return_book, width=20, size=14).place(x=menu_x, y=560, width=260, height=48)
        self.make_button(canvas, "Lich su muon", self.show_activity, width=20, size=14).place(x=menu_x, y=625, width=260, height=48)

        sw = self.root.winfo_screenwidth()
        self.make_button(canvas, "Dang xuat", self.show_login, width=14, size=12, bg=COLOR_ACCENT, fg="white").place(x=sw-220, y=28, width=170, height=44)
        self.work_area = canvas

    def create_tree(self, parent, columns):
        # Tang kich thuoc cac cot de hien day du ten sach va bang rong ve ben phai
        column_widths = {
            "Ma sach": 90,
            "Ten sach": 280,
            "Tac gia": 180,
            "The loai": 190,
            "So luong": 90,
            "Vi tri": 100,
            "Ma sinh vien": 130,
            "Ngay muon": 130,
            "Ngay tra": 130,
        }

        tree_frame = Frame(parent, bg=COLOR_PANEL)
        tree_frame.pack(fill=BOTH, expand=True, padx=15, pady=10)

        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=13)

        y_scroll = Scrollbar(tree_frame, orient=VERTICAL, command=tree.yview)
        x_scroll = Scrollbar(tree_frame, orient=HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        y_scroll.pack(side=RIGHT, fill=Y)
        x_scroll.pack(side=BOTTOM, fill=X)
        tree.pack(fill=BOTH, expand=True)

        for col in columns:
            width = column_widths.get(col, 120)
            anchor = "e" if col in ("So luong",) else "w"
            tree.heading(col, text=col, anchor="center")
            tree.column(col, width=width, minwidth=width, anchor=anchor, stretch=True)
        return tree

    def show_add_book(self):
        panel = self.make_panel(self.work_area)
        Label(panel, text="THEM SACH", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 18, "bold")).pack(pady=15)

        self.book_id_var = StringVar()
        self.title_var = StringVar()
        self.author_var = StringVar()
        self.genre_var = StringVar()
        self.copies_var = IntVar(value=1)
        self.location_var = StringVar()

        fields = [
            ("Ma sach", self.book_id_var),
            ("Ten sach", self.title_var),
            ("Tac gia", self.author_var),
            ("The loai", self.genre_var),
            ("So luong", self.copies_var),
            ("Vi tri", self.location_var),
        ]
        for label, var in fields:
            row = Frame(panel, bg=COLOR_PANEL)
            row.pack(pady=6)
            Label(row, text=label, width=15, anchor="w", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 12, "bold")).pack(side="left")
            Entry(row, textvariable=var, width=35, bg=COLOR_ACCENT, fg="white", font=("Arial", 12)).pack(side="left")

        Button(panel, text="Them", bg=COLOR_ACCENT, fg="white", font=("Arial", 11, "bold"), width=14, command=self.add_book).pack(pady=20)

    def add_book(self):
        try:
            book = Book(
                self.book_id_var.get().strip(),
                self.title_var.get().strip(),
                self.author_var.get().strip(),
                self.genre_var.get().strip(),
                int(self.copies_var.get()),
                self.location_var.get().strip(),
            )
            if not book.is_valid():
                messagebox.showwarning("Thong bao", "Thong tin sach khong hop le hoac so luong bi am")
                return
            self.db.add_book(book)
            messagebox.showinfo("Thanh cong", "Them sach thanh cong")
        except sqlite3.IntegrityError:
            messagebox.showerror("Loi", "Ma sach da ton tai")
        except ValueError:
            messagebox.showerror("Loi", "So luong phai la so nguyen")

    def show_search_book(self):
        panel = self.make_panel(self.work_area, width=1000, height=570, x=320, y=70)
        Label(panel, text="TIM KIEM SACH", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 18, "bold")).pack(pady=10)
        self.search_var = StringVar()
        row = Frame(panel, bg=COLOR_PANEL)
        row.pack(pady=8)
        Entry(row, textvariable=self.search_var, width=35, bg=COLOR_ACCENT, fg="white", font=("Arial", 12)).pack(side="left", padx=5)
        Button(row, text="Tim", bg=COLOR_ACCENT, fg="white", font=("Arial", 11, "bold"), command=self.search_book).pack(side="left", padx=5)

        self.book_tree = self.create_tree(panel, ("Ma sach", "Ten sach", "Tac gia", "The loai", "So luong", "Vi tri"))

        action = Frame(panel, bg=COLOR_PANEL)
        action.pack(pady=10)
        self.number_var = IntVar(value=1)
        Entry(action, textvariable=self.number_var, width=8, bg=COLOR_ACCENT, fg="white", font=("Arial", 11)).pack(side="left", padx=5)
        Button(action, text="Them so luong", bg=COLOR_ACCENT, fg="white", command=self.add_copies).pack(side="left", padx=5)
        Button(action, text="Giam so luong", bg=COLOR_ACCENT, fg="white", command=self.delete_copies).pack(side="left", padx=5)
        Button(action, text="Xoa sach", bg=COLOR_ACCENT, fg="white", command=self.delete_book).pack(side="left", padx=5)

    def fill_tree(self, tree, rows):
        for item in tree.get_children():
            tree.delete(item)
        for row in rows:
            tree.insert("", "end", values=row)

    def search_book(self):
        keyword = self.search_var.get().strip()
        if not keyword:
            messagebox.showwarning("Thong bao", "Vui long nhap tu khoa")
            return
        rows = self.db.search_books(keyword)
        self.fill_tree(self.book_tree, rows)
        if not rows:
            messagebox.showinfo("Thong bao", "Khong tim thay du lieu")

    def selected_book_id(self):
        selected = self.book_tree.focus()
        if not selected:
            messagebox.showwarning("Thong bao", "Vui long chon mot dong trong bang")
            return None
        return self.book_tree.item(selected, "values")[0]

    def add_copies(self):
        book_id = self.selected_book_id()
        if not book_id:
            return
        number = int(self.number_var.get())
        if number < 0:
            messagebox.showwarning("Thong bao", "So luong khong duoc am")
            return
        self.db.add_copies(book_id, number)
        messagebox.showinfo("Thanh cong", "Da them so luong sach")
        self.search_book()

    def delete_copies(self):
        book_id = self.selected_book_id()
        if not book_id:
            return
        try:
            number = int(self.number_var.get())
            if number < 0:
                messagebox.showwarning("Thong bao", "So luong khong duoc am")
                return
            self.db.delete_copies(book_id, number)
            messagebox.showinfo("Thanh cong", "Da giam so luong sach")
            self.search_book()
        except ValueError as e:
            messagebox.showerror("Loi", str(e))

    def delete_book(self):
        book_id = self.selected_book_id()
        if not book_id:
            return
        try:
            self.db.delete_book(book_id)
            messagebox.showinfo("Thanh cong", "Da xoa sach")
            self.search_book()
        except ValueError as e:
            messagebox.showerror("Loi", str(e))

    def import_books_excel(self):
        file_path = filedialog.askopenfilename(
            title="Chon file Excel danh sach sach",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not file_path:
            return

        try:
            added, skipped = self.db.import_books_from_excel(file_path)
            messagebox.showinfo(
                "Ket qua",
                f"Doc file Excel thanh cong!\nThem moi: {added} sach\nBo qua: {skipped} dong"
            )
        except ImportError as e:
            messagebox.showerror("Thieu thu vien", str(e))
        except Exception as e:
            messagebox.showerror("Loi", f"Khong doc duoc file Excel: {e}")

    def show_all_books(self):
        panel = self.make_panel(self.work_area, width=1000, height=520, x=320, y=90)
        Label(panel, text="DANH SACH SACH", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 18, "bold")).pack(pady=15)
        tree = self.create_tree(panel, ("Ma sach", "Ten sach", "Tac gia", "The loai", "So luong", "Vi tri"))
        self.fill_tree(tree, self.db.get_all_books())

    def show_issue_book(self):
        panel = self.make_panel(self.work_area, width=520, height=350, x=520, y=130)
        Label(panel, text="MUON SACH", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 18, "bold")).pack(pady=20)
        self.issue_book_var = StringVar()
        self.issue_student_var = StringVar()
        self.form_row(panel, "Ma sach", self.issue_book_var)
        self.form_row(panel, "Ma sinh vien", self.issue_student_var)
        Button(panel, text="Muon sach", bg=COLOR_ACCENT, fg="white", font=("Arial", 11, "bold"), width=16, command=self.issue_book).pack(pady=20)

    def form_row(self, panel, label, variable):
        row = Frame(panel, bg=COLOR_PANEL)
        row.pack(pady=8)
        Label(row, text=label, width=14, anchor="w", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 12, "bold")).pack(side="left")
        Entry(row, textvariable=variable, width=28, bg=COLOR_ACCENT, fg="white", font=("Arial", 12)).pack(side="left")

    def issue_book(self):
        book_id = self.issue_book_var.get().strip()
        student_id = self.issue_student_var.get().strip()
        if not book_id or not student_id:
            messagebox.showwarning("Thong bao", "Vui long nhap day du thong tin")
            return
        try:
            self.db.issue_book(BorrowRecord(book_id, student_id))
            messagebox.showinfo("Thanh cong", "Muon sach thanh cong")
        except sqlite3.IntegrityError:
            messagebox.showerror("Loi", "Sinh vien nay da muon quyen sach nay")
        except ValueError as e:
            messagebox.showerror("Loi", str(e))

    def show_return_book(self):
        panel = self.make_panel(self.work_area, width=520, height=350, x=520, y=130)
        Label(panel, text="TRA SACH", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 18, "bold")).pack(pady=20)
        self.return_book_var = StringVar()
        self.return_student_var = StringVar()
        self.form_row(panel, "Ma sach", self.return_book_var)
        self.form_row(panel, "Ma sinh vien", self.return_student_var)
        Button(panel, text="Tra sach", bg=COLOR_ACCENT, fg="white", font=("Arial", 11, "bold"), width=16, command=self.return_book).pack(pady=20)

    def return_book(self):
        book_id = self.return_book_var.get().strip()
        student_id = self.return_student_var.get().strip()
        if not book_id or not student_id:
            messagebox.showwarning("Thong bao", "Vui long nhap day du thong tin")
            return
        try:
            self.db.return_book(book_id, student_id)
            messagebox.showinfo("Thanh cong", "Tra sach thanh cong")
        except ValueError as e:
            messagebox.showerror("Loi", str(e))

    def show_activity(self):
        panel = self.make_panel(self.work_area, width=820, height=520, x=430, y=90)
        Label(panel, text="LICH SU MUON SACH", bg=COLOR_PANEL, fg=COLOR_ACCENT_HOVER, font=("Arial", 18, "bold")).pack(pady=12)
        self.activity_var = StringVar()
        row = Frame(panel, bg=COLOR_PANEL)
        row.pack(pady=5)
        Entry(row, textvariable=self.activity_var, width=30, bg=COLOR_ACCENT, fg="white", font=("Arial", 12)).pack(side="left", padx=5)
        Button(row, text="Tim", bg=COLOR_ACCENT, fg="white", command=self.search_activity).pack(side="left", padx=5)
        Button(row, text="Tat ca", bg=COLOR_ACCENT, fg="white", command=self.show_all_activity).pack(side="left", padx=5)
        self.activity_tree = self.create_tree(panel, ("Ma sach", "Ma sinh vien", "Ngay muon", "Ngay tra"))

    def search_activity(self):
        keyword = self.activity_var.get().strip()
        if not keyword:
            messagebox.showwarning("Thong bao", "Vui long nhap ma sach hoac ma sinh vien")
            return
        rows = self.db.search_activity(keyword)
        self.fill_tree(self.activity_tree, rows)
        if not rows:
            messagebox.showinfo("Thong bao", "Khong tim thay du lieu")

    def show_all_activity(self):
        self.fill_tree(self.activity_tree, self.db.get_all_activity())


if __name__ == "__main__":
    LibraryApp()
   